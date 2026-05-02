import streamlit as st
import pdfplumber
import pandas as pd
import google.generativeai as genai
import io
import json
import google.api_core.exceptions
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.chart import BarChart, Reference

API_KEY = st.secrets["GEMINI_API_KEY"]
genai.configure(api_key=API_KEY)

def extract_full_text(file):
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"
    return text

st.set_page_config(page_title="Andalusia-Evaluator", page_icon="tab.svg", layout="wide")

st.markdown("""
<style>
    .stApp {
        background-color: #FDFBF7;
    }
    header {
        background-color: transparent !important;
    }
    h1, h2, h3 {
        color: #004B87 !important;
    }
    .stButton>button {
        background-color: #004B87;
        color: white;
        border-radius: 8px;
        font-weight: bold;
        border: none;
        padding: 10px 20px;
        width: 100%;
    }
    .stButton>button:hover {
        background-color: #003366;
        color: white;
    }
    hr {
        border-top: 2px solid #D9E2EC;
    }
</style>
""", unsafe_allow_html=True)

col_logo, col_title = st.columns([1.5, 4])

with col_logo:
    st.image("FooterLogo-eg.svg", use_container_width=True) 

with col_title:
    st.markdown("<h1 style='margin-top: -15px;'>Medical Equipment Evaluation System</h1>", unsafe_allow_html=True)

st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    std_file = st.file_uploader("📑 Upload Standard Specs (PDF)", type="pdf")
with col2:
    offer_files = st.file_uploader("🏢 Upload Vendor Offers (PDF)", type="pdf", accept_multiple_files=True)

if st.button("🚀 Analyze & Generate Ultimate Excel"):
    if std_file and offer_files:
        
        model = genai.GenerativeModel(
            'gemini-2.5-flash-lite',
            generation_config={"response_mime_type": "application/json"}
        )
        
        with st.spinner("Step 1: Extracting Master Standard Template..."):
            std_text = extract_full_text(std_file)
            std_prompt = f"""
            Extract all technical specifications from this Hospital Standard document.
            Return STRICTLY a JSON array of objects with exactly these keys:
            "Feature": "Name of the parameter",
            "Standard": "Required value"
            Document: {std_text}
            """
            try:
                std_response = model.generate_content(std_prompt)
                standard_template = json.loads(std_response.text)
                template_str = json.dumps(standard_template, ensure_ascii=False)
            except google.api_core.exceptions.ResourceExhausted:
                st.error("⚠️ Trial Limit Exceeded: The daily free quota for the AI has run out. Please try again tomorrow.")
                st.stop()
            except Exception:
                st.error("⚠️ An error occurred while processing the Standard PDF. Please ensure the file is readable.")
                st.stop()

        with st.spinner("Step 2: Evaluating Vendor Offers..."):
            main_df = None
            company_names = []
            summaries = []

            for offer in offer_files:
                offer_text = extract_full_text(offer)
                prompt = f"""
                You are a Senior Biomedical Engineer. Evaluate the Vendor Offer against the EXACT provided Hospital Standard Template.
                Hospital Standard Template (JSON):
                {template_str}
                Vendor Offer Text:
                {offer_text}
                Output STRICTLY a JSON object with this exact structure:
                {{
                  "Company_Name": "Extract Vendor/Company name",
                  "Final_Recommendation": "Accepted or Rejected",
                  "Overall_Score": <number out of 10>,
                  "Final_Reason": "English justification",
                  "Specifications": [
                    {{
                      "Feature": "MUST BE EXACTLY THE SAME AS IN THE TEMPLATE",
                      "Standard": "MUST BE EXACTLY THE SAME AS IN THE TEMPLATE",
                      "Vendor_Value": "Offered value from text (or 'Not Mentioned')",
                      "Match": "Yes or No",
                      "Grade": <number out of 10>,
                      "Comment": "English technical comment"
                    }}
                  ]
                }}
                """
                
                try:
                    response = model.generate_content(prompt)
                    data = json.loads(response.text)
                    comp_name = data.get("Company_Name", "Unknown Vendor")
                    company_names.append(comp_name)
                    
                    summaries.append({
                        "Company": comp_name,
                        "Final Recommendation": data.get("Final_Recommendation", ""),
                        "Overall Score": data.get("Overall_Score", 0),
                        "Reason": data.get("Final_Reason", "")
                    })
                    
                    df = pd.DataFrame(data["Specifications"])
                    cols = {
                        'Vendor_Value': f'{comp_name} Offer',
                        'Match': f'{comp_name} Match',
                        'Grade': f'{comp_name} Grade',
                        'Comment': f'{comp_name} Comment'
                    }
                    df.rename(columns=cols, inplace=True)
                    
                    if main_df is None:
                        main_df = df
                    else:
                        main_df = pd.merge(main_df, df, on=['Feature', 'Standard'], how='left')
                        
                    st.success(f"✅ Analyzed: {comp_name}")
                
                except google.api_core.exceptions.ResourceExhausted:
                    st.error(f"⚠️ Trial Limit Exceeded while analyzing '{offer.name}'. Please try again tomorrow.")
                    break
                except Exception:
                    st.error(f"⚠️ An unexpected error occurred while analyzing '{offer.name}'. Skipping this file.")
                    continue

            if main_df is not None and not main_df.empty:
                ordered_cols = ['Feature', 'Standard']
                for i, comp in enumerate(company_names):
                    if i > 0:
                        sep_col = f'   _{i}'
                        main_df[sep_col] = ""
                        ordered_cols.append(sep_col)
                    
                    expected_cols = [f'{comp} Offer', f'{comp} Match', f'{comp} Grade', f'{comp} Comment']
                    for col in expected_cols:
                        if col not in main_df.columns:
                            main_df[col] = ""
                    ordered_cols.extend(expected_cols)
                
                main_df = main_df[ordered_cols]

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    main_df.to_excel(writer, index=False, sheet_name='Detailed Comparison')
                    workbook = writer.book
                    worksheet = writer.sheets['Detailed Comparison']
                    
                    base_font = Font(size=14)
                    header_fill = PatternFill(start_color='004B87', end_color='004B87', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True, size=14)
                    
                    row_even_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid') 
                    row_odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  
                    
                    yes_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
                    no_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
                    yes_no_font = Font(color='FFFFFF', bold=True, size=14)

                    thin_border = Border(left=Side(style='thin', color='D9E2EC'),
                                         right=Side(style='thin', color='D9E2EC'),
                                         top=Side(style='thin', color='D9E2EC'),
                                         bottom=Side(style='thin', color='D9E2EC'))
                                         
                    gray_fill = PatternFill(start_color='E4E7EB', end_color='E4E7EB', fill_type='solid')

                    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
                        for cell in row:
                            col_letter = cell.column_letter
                            col_name = str(worksheet.cell(row=1, column=cell.column).value)
                            
                            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                            cell.border = thin_border
                            cell.font = base_font
                            
                            if '   _' in col_name:
                                worksheet.column_dimensions[col_letter].width = 4
                                cell.fill = gray_fill
                                continue
                            else:
                                worksheet.column_dimensions[col_letter].width = 35
                                
                            if cell.row == 1:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                            else:
                                if cell.row % 2 == 0:
                                    cell.fill = row_even_fill
                                else:
                                    cell.fill = row_odd_fill

                                val = str(cell.value).strip().lower()
                                if val == 'yes':
                                    cell.fill = yes_fill
                                    cell.font = yes_no_font
                                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                                elif val == 'no':
                                    cell.fill = no_fill
                                    cell.font = yes_no_font
                                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                    df_summary = pd.DataFrame(summaries)
                    df_summary_t = df_summary.set_index('Company').T.reset_index()
                    df_summary_t.rename(columns={'index': 'Metric'}, inplace=True)
                    
                    df_summary_t.to_excel(writer, index=False, sheet_name='Summary & Chart')
                    ws_summary = writer.sheets['Summary & Chart']
                    
                    for row in ws_summary.iter_rows(min_row=1, max_col=ws_summary.max_column, max_row=ws_summary.max_row):
                        for cell in row:
                            cell.border = thin_border
                            cell.font = base_font
                            if cell.row == 1:
                                cell.fill = header_fill
                                cell.font = header_font
                            ws_summary.column_dimensions[cell.column_letter].width = 45
                            cell.alignment = Alignment(wrap_text=True, vertical='center')
                    
                    if len(summaries) > 0:
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = "Evaluation Scores"
                        chart.y_axis.title = "Score (10)"
                        chart.x_axis.title = "Vendor"
                        
                        data_ref = Reference(ws_summary, min_col=1, min_row=3, max_col=len(summaries)+1, max_row=3)
                        cats_ref = Reference(ws_summary, min_col=2, min_row=1, max_col=len(summaries)+1)
                        
                        chart.add_data(data_ref, from_rows=True, titles_from_data=True)
                        chart.set_categories(cats_ref)
                        chart.height = 15
                        chart.width = 20
                        ws_summary.add_chart(chart, "A6")

                dynamic_filename = "Andalusia_Comparison_" + "_vs_".join(company_names) + ".xlsx"
                st.download_button(
                    label=f"📥 Download Professional Report",
                    data=output.getvalue(),
                    file_name=dynamic_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.warning("Please upload the Standard PDF and Vendor Offers.")
